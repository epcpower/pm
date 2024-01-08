import attr
import openpyxl
import typing
import pathlib

import epyqlib.pm.parametermodel
import epyqlib.utils.general

import epcpm.c
import epcpm.pm_helper
import epcpm.anomalymodel

builders = epyqlib.utils.general.TypeMap()


@attr.s
class Fields(epcpm.pm_helper.FieldsInterface):
    """The fields defined for a given row in the output XLS file."""

    group = attr.ib(default=None, type=typing.Union[str, bool])
    name = attr.ib(default=None, type=typing.Union[str, bool])
    code = attr.ib(default=None, type=typing.Union[int, bool])
    trig_type = attr.ib(default=None, type=typing.Union[int, bool])
    response_level_inactive = attr.ib(default=None, type=typing.Union[int, bool])
    response_level_active = attr.ib(default=None, type=typing.Union[int, bool])
    comment = attr.ib(default=None, type=typing.Union[str, bool])


field_names = Fields(
    group="Group",
    name="Name",
    code="Code",
    trig_type="Trigger type",
    response_level_inactive="Response level inactive",
    response_level_active="Response level active",
    comment="Description",
)


@attr.s
class InfoSheetFields(epcpm.pm_helper.FieldsInterface):
    """The fields defined for an enumerator information sheet."""

    name = attr.ib(default=None, type=typing.Union[str, bool])
    desc = attr.ib(default=None, type=typing.Union[str, bool])


info_sheet_field_names = InfoSheetFields(
    name="Name",
    desc="Description",
)


def export(
    path: pathlib.Path,
    anomaly_model: epyqlib.attrsmodel.Model,
    parameters_model: epyqlib.attrsmodel.Model,
    column_filter: [epcpm.pm_helper.FieldsInterface, None] = None,
    skip_output: bool = False,
) -> None:
    """
    Exports anomaly information to a spreadsheet file.

    Args:
        path:             Path to output XLSX file.
        anomaly_model:    Anomaly data model.
        parameters_model: Parameter data model.
        column_filter:    Optional filter to remove selected output columns.
        skip_output:      Optional boolean flag to disable output.

    Returns:
        None
    """

    if column_filter is None:
        column_filter = epcpm.pm_helper.attr_fill(Fields, True)

    builder = epcpm.anomaliestoxlsx.builders.wrap(
        wrapped=anomaly_model.root,
        column_filter=column_filter,
        parameter_uuid_finder=parameters_model.node_from_uuid,
        response_levels=anomaly_model.list_selection_roots["anomaly_response_levels"],
        trigger_types=anomaly_model.list_selection_roots["anomaly_trigger_types"],
        skip_output=skip_output,
    )

    workbook = builder.gen()
    if workbook:
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)


@builders(epcpm.anomalymodel.Root)
@attr.s
class Root:
    wrapped = attr.ib()
    column_filter = attr.ib()
    parameter_uuid_finder = attr.ib()
    response_levels = attr.ib()
    trigger_types = attr.ib()
    skip_output = attr.ib(default=False)

    header_font = openpyxl.styles.Font(bold=True)
    header_fill = openpyxl.styles.PatternFill(
        start_color="FF8DB4E2", end_color="FF8DB4E2", fill_type="solid"
    )

    def gen(self):

        if self.skip_output:
            return None

        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)

        # Add sheet for response level descriptions
        worksheet = workbook.create_sheet("Anomalies")
        self.generate_anomalies_sheet(worksheet)

        # Add sheet for response level descriptions
        worksheet = workbook.create_sheet("Response levels")
        self.generate_enum_info_sheet(worksheet, self.response_levels)

        # Add sheet for trigger type information
        worksheet = workbook.create_sheet("Trigger types")
        self.generate_enum_info_sheet(worksheet, self.trigger_types)

        return workbook

    def generate_anomalies_sheet(
        self, worksheet: openpyxl.worksheet.worksheet.Worksheet
    ) -> None:
        """
        Generates a spreadsheet for anomalies.

        Args:
            worksheet: Empty openpyxl worksheet object.

        Returns:
            None
        """

        # Helper function for converting cell value to string
        def as_text(value):
            if value is None:
                return ""
            return str(value)

        # Add and format header row
        worksheet.append(field_names.as_filtered_tuple(self.column_filter))
        for cell in worksheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Iterate anomaly tables and add new rows from them
        for anomaly_table in self.wrapped.children:

            rows = builders.wrap(
                wrapped=anomaly_table,
                parameter_uuid_finder=self.parameter_uuid_finder,
            ).gen()

            for row in rows:
                worksheet.append(
                    row.as_filtered_tuple(self.column_filter),
                )

        # Adjust column widths in the worksheet in regards of text length
        for column_cells in worksheet.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column].width = length + 5

    def generate_enum_info_sheet(
        self,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        enumeration: epyqlib.pm.parametermodel.Enumeration,
    ) -> None:
        """
        Generates an enumeration information spread sheet with two columns.
        The sheet contains enumerator names and their respective comments.

        Args:
            worksheet:   Empty openpyxl worksheet object.
            enumeration: Enumeration for which the sheet is generated.

        Returns:
            None
        """

        # Add and format header row
        worksheet.append(
            info_sheet_field_names.as_filtered_tuple(self.column_filter),
        )
        for cell in worksheet[1]:
            cell.font = self.header_font
            cell.fill = self.header_fill

        # Add response level descriptions
        row = InfoSheetFields()
        for enum in enumeration.children:
            row.name = enum.name
            row.desc = enum.description
            worksheet.append(
                row.as_filtered_tuple(self.column_filter),
            )

        # Format cell text alignment for all cells in the sheet
        for col in worksheet.iter_cols():
            for cell in col:
                cell.alignment = openpyxl.styles.Alignment(
                    wrap_text=True, horizontal="left", vertical="top"
                )

        # Format column widths
        worksheet.column_dimensions["A"].width = 50
        worksheet.column_dimensions["B"].width = 200

        # Format row heights
        for cell in worksheet["A"][1:]:
            worksheet.row_dimensions[cell.row].height = 50


@builders(epcpm.anomalymodel.AnomalyTable)
@attr.s
class AnomalyTable:

    wrapped = attr.ib()
    parameter_uuid_finder = attr.ib()

    def gen(self):
        anomalies = list(self.wrapped.children)

        rows = []
        for anomaly in anomalies:
            row = builders.wrap(
                wrapped=anomaly, parameter_uuid_finder=self.parameter_uuid_finder
            ).gen()
            row.group = self.wrapped.name
            rows.append(row)
        return rows


@builders(epcpm.anomalymodel.Anomaly)
@attr.s
class Anomaly:
    wrapped = attr.ib()
    parameter_uuid_finder = attr.ib()

    def gen(self):
        row = Fields()

        # Resolve response level and trigger type names
        response_level_active = (
            self.parameter_uuid_finder(self.wrapped.response_level_active).name
            if self.wrapped.response_level_active
            else None
        )
        response_level_inactive = (
            self.parameter_uuid_finder(self.wrapped.response_level_inactive).name
            if self.wrapped.response_level_inactive
            else None
        )
        trig_type = (
            self.parameter_uuid_finder(self.wrapped.trigger_type).name
            if self.wrapped.trigger_type
            else None
        )

        # Set row values
        row.name = self.wrapped.name
        row.code = self.wrapped.code
        row.trig_type = trig_type
        row.response_level_active = response_level_active
        row.response_level_inactive = response_level_inactive
        row.comment = self.wrapped.comment

        return row
