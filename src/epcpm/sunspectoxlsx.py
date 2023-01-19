from __future__ import (
    annotations,
)  # See PEP 563, check to remove in future Python version higher than 3.7
import itertools
import math

import attr
import openpyxl
import typing
import uuid

import epyqlib.pm.parametermodel
import epyqlib.utils.general

import epcpm.c
import epcpm.pm_helper
import epcpm.sunspecmodel
import epcpm.sunspectointerface

from enum import Enum

builders = epyqlib.utils.general.TypeMap()
enumeration_builders = epyqlib.utils.general.TypeMap()
enumerator_builders = epyqlib.utils.general.TypeMap()


data_point_fields = attr.fields(epcpm.sunspecmodel.DataPoint)
epc_enumerator_fields = attr.fields(
    epyqlib.pm.parametermodel.SunSpecEnumerator,
)
bitfield_fields = attr.fields(epcpm.sunspecmodel.DataPointBitfield)


class DummyModels(Enum):
    LOWER_BOUND = 1000
    UPPER_BOUND = 1999

    @staticmethod
    def within_bounds(value):
        return DummyModels.LOWER_BOUND.value <= value <= DummyModels.UPPER_BOUND.value


@attr.s
class Fields(epcpm.pm_helper.FieldsInterface):
    """The fields defined for a given row in the output XLS file."""

    field_type = attr.ib(default=None, type=typing.Union[str, bool])
    applicable_point = attr.ib(default=None, type=typing.Union[str, bool])
    address_offset = attr.ib(default=None, type=typing.Union[str, bool, int])
    block_offset = attr.ib(default=None, type=typing.Union[str, bool, int])
    size = attr.ib(default=None, type=typing.Union[str, bool])
    name = attr.ib(default=None, type=typing.Union[str, bool])
    label = attr.ib(default=None, type=typing.Union[str, bool])
    value = attr.ib(default=None, type=typing.Union[str, bool, int])
    type = attr.ib(default=None, type=typing.Union[str, bool])
    units = attr.ib(default=None, type=typing.Union[str, bool])
    scale_factor = attr.ib(default=None, type=typing.Union[str, bool])
    read_write = attr.ib(default=None, type=typing.Union[str, bool])
    mandatory = attr.ib(default=None, type=typing.Union[str, bool])
    description = attr.ib(default=None, type=typing.Union[str, bool])
    notes = attr.ib(default=None, type=typing.Union[str, bool])
    modbus_address = attr.ib(default=None, type=typing.Union[str, bool, int])
    get = attr.ib(default=None, type=typing.Union[str, bool])
    set = attr.ib(default=None, type=typing.Union[str, bool])
    item = attr.ib(default=None, type=typing.Union[str, bool])


field_names = Fields(
    field_type="Field Type",
    applicable_point="Applicable Point",
    address_offset="Address Offset",
    block_offset="Block Offset",
    size="Size",
    name="Name",
    label="Label",
    value="Value",
    type="Type",
    units="Units",
    scale_factor="SF",
    read_write="R/W",
    mandatory="Mandatory M/O",
    description="Description",
    notes="Notes",
    modbus_address="Modbus Address",
    get="get",
    set="set",
    item="item",
)


point_fields = Fields(
    block_offset=data_point_fields.block_offset,
    size=data_point_fields.size,
    type=data_point_fields.type_uuid,
    scale_factor=data_point_fields.factor_uuid,
    units=data_point_fields.units,
)

bitfield_fields = Fields(
    block_offset=bitfield_fields.block_offset,
    size=bitfield_fields.size,
    type=bitfield_fields.type_uuid,
)

enumerator_fields = Fields(
    name=epc_enumerator_fields.abbreviation,
    label=epc_enumerator_fields.label,
    description=epc_enumerator_fields.description,
    notes=epc_enumerator_fields.notes,
    value=epc_enumerator_fields.value,
    field_type=epc_enumerator_fields.type,
)


def export(
    path,
    sunspec_model,
    sunspec_id,
    parameters_model,
    column_filter=None,
    skip_sunspec=False,
    output_dummy_models=True,
):
    if column_filter is None:
        column_filter = epcpm.pm_helper.attr_fill(Fields, True)

    builder = epcpm.sunspectoxlsx.builders.wrap(
        wrapped=sunspec_model.root,
        parameter_uuid_finder=sunspec_model.node_from_uuid,
        parameter_model=parameters_model,
        sunspec_id=sunspec_id,
        skip_sunspec=skip_sunspec,
        column_filter=column_filter,
        output_dummy_models=output_dummy_models,
    )

    workbook = builder.gen()

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


@builders(epcpm.sunspecmodel.Root)
@attr.s
class Root:
    wrapped = attr.ib()
    column_filter = attr.ib()
    skip_sunspec = attr.ib(default=False)
    parameter_uuid_finder = attr.ib(default=None)
    parameter_model = attr.ib(default=None)
    sunspec_id = attr.ib(default=None)
    sort_models = attr.ib(default=False)
    output_dummy_models = attr.ib(default=True)

    def gen(self):
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)

        workbook.create_sheet("License Agreement")
        workbook.create_sheet("Summary")
        workbook.create_sheet("Index")

        if not self.skip_sunspec:
            if self.sort_models:
                children = sorted(
                    self.wrapped.children,
                    key=lambda child: (
                        0 if isinstance(child, epcpm.sunspecmodel.Model) else 1,
                        getattr(child, "id", math.inf),
                    ),
                )
            else:
                children = list(self.wrapped.children)

            # Account for 'SunS' length.
            model_offset = epcpm.pm_helper.SUNS_LENGTH
            for model in children:
                if isinstance(model, epcpm.sunspecmodel.Table):
                    # TODO: for now, implement it soon...
                    continue

                worksheet = workbook.create_sheet()

                model_offset += builders.wrap(
                    wrapped=model,
                    worksheet=worksheet,
                    padding_type=self.parameter_model.list_selection_roots[
                        "sunspec types"
                    ].child_by_name("pad"),
                    parameter_uuid_finder=self.parameter_uuid_finder,
                    sunspec_id=self.sunspec_id,
                    column_filter=self.column_filter,
                    model_offset=model_offset,
                ).gen()

                if not self.output_dummy_models and DummyModels.within_bounds(model.id):
                    # Do not put dummy models in the spreadsheet.
                    workbook.remove(worksheet)

        return workbook


@builders(epcpm.sunspecmodel.Model)
@attr.s
class Model:
    wrapped = attr.ib()
    worksheet = attr.ib()
    padding_type = attr.ib()
    column_filter = attr.ib()
    model_offset = attr.ib()  # starting Modbus address for the model
    parameter_uuid_finder = attr.ib(default=None)
    sunspec_id = attr.ib(default=None)

    def gen(self):
        self.worksheet.title = str(self.wrapped.id)
        self.worksheet.append(field_names.as_filtered_tuple(self.column_filter))

        self.wrapped.children[0].check_offsets_and_length()

        non_header_length = 0
        overall_length = 0
        accumulated_length = 0  # Used for each point row on spreadsheet
        rows = []

        # Discover the FixedBlock for future use in the TableBlock.
        for block in self.wrapped.children:
            if isinstance(block, epcpm.sunspecmodel.FixedBlock):
                fixed_block_reference = block
                break
        else:
            fixed_block_reference = None

        model_types = ["Header", "Fixed Block", "Repeating Block"]
        child_model_types = zip(enumerate(self.wrapped.children), model_types)
        for (i, child), model_type in child_model_types:
            add_padding = epcpm.pm_helper.add_padding_to_block(
                child, self.sunspec_id, self.wrapped.id, model_type
            )

            builder = builders.wrap(
                wrapped=child,
                add_padding=add_padding,
                padding_type=self.padding_type,
                model_type=model_type,
                model_id=self.wrapped.id,
                model_offset=self.model_offset,
                parameter_uuid_finder=self.parameter_uuid_finder,
                sunspec_id=self.sunspec_id,
                address_offset=accumulated_length,
                is_table=model_type == "Repeating Block",
                fixed_block_reference=fixed_block_reference,
            )

            built_rows, block_length = builder.gen()
            accumulated_length += block_length
            if len(built_rows) > 0:
                rows.extend(built_rows)
                rows.append(Fields())

            # The code keeping track of overall_length, non_header_length, and accumulated_length
            # could be refactored, but better to keep it simple so that the logic and purpose is clear.
            # Each of the values has a unique purpose.
            if i == 0:
                # For HeaderBlock
                overall_length += block_length
            elif i == 1:
                # For FixedBlock
                overall_length += block_length
                non_header_length += block_length
            elif i == 2:
                # For TableRepeatingBlock only
                overall_length += block_length * child.get_num_repeats()
                non_header_length += block_length * child.get_num_repeats()

        for i, row in enumerate(rows):
            if i == 0:
                row.value = self.wrapped.id
            elif i == 1:
                row.value = non_header_length

            self.worksheet.append(row.as_filtered_tuple(self.column_filter))

        for block in self.wrapped.children:
            builder = enumeration_builders.wrap(
                wrapped=block,
                parameter_uuid_finder=self.parameter_uuid_finder,
            )
            rows = builder.gen()

            for row in rows:
                self.worksheet.append(
                    row.as_filtered_tuple(self.column_filter),
                )

        return overall_length


@builders(epcpm.sunspecmodel.Table)
@attr.s
class Table:
    wrapped = attr.ib()
    worksheet = attr.ib()
    padding_type = attr.ib()
    column_filter = attr.ib()
    parameter_uuid_finder = attr.ib(default=None)

    def gen(self):
        return []


@enumeration_builders(epcpm.sunspecmodel.DataPoint)
@attr.s
class Enumeration:
    wrapped = attr.ib()
    point = attr.ib()
    parameter_uuid_finder = attr.ib(default=None)

    def gen(self):
        rows = []

        enumeration_uuid = getattr(self.wrapped, "enumeration_uuid", None)
        if enumeration_uuid is None:
            return rows

        enumeration = self.parameter_uuid_finder(enumeration_uuid)

        enumerators_by_bit = {
            enumerator.value: enumerator for enumerator in enumeration.children
        }

        # 16 bits per register
        total_bit_count = self.wrapped.size * 16
        decimal_digits = len(str(total_bit_count - 1))

        for bit in range(total_bit_count):
            enumerator = enumerators_by_bit.get(bit)
            if enumerator is None:
                padded_bit_string = f"{bit:0{decimal_digits}}"
                enumerator = epyqlib.pm.parametermodel.SunSpecEnumerator(
                    label=f"Reserved - {padded_bit_string}",
                    name=f"Rsvd{padded_bit_string}",
                    value=bit,
                )

            builder = enumerator_builders.wrap(
                wrapped=enumerator,
                point=self.point,
                parameter_uuid_finder=self.parameter_uuid_finder,
            )
            rows.append(builder.gen())

        return rows


@enumerator_builders(epyqlib.pm.parametermodel.SunSpecEnumerator)
@attr.s
class Enumerator:
    wrapped = attr.ib()
    point = attr.ib()
    parameter_uuid_finder = attr.ib(default=None)

    # TODO: CAMPid 07397546759269756456100183066795496952476951653
    def gen(self):
        row = Fields()

        for name, field in attr.asdict(enumerator_fields).items():
            if field is None:
                continue

            setattr(row, name, getattr(self.wrapped, field.name))

        if row.name is None:
            row.name = self.wrapped.name

        parameter = self.parameter_uuid_finder(self.point.parameter_uuid)
        row.applicable_point = parameter.abbreviation

        field_type_parameter = self.parameter_uuid_finder(self.point.type_uuid)
        row.field_type = field_type_parameter.name

        return row


@builders(epcpm.sunspecmodel.TableRepeatingBlock)
@builders(epcpm.sunspecmodel.HeaderBlock)
@builders(epcpm.sunspecmodel.FixedBlock)
@attr.s
class Block:
    wrapped = attr.ib()
    add_padding = attr.ib()
    padding_type = attr.ib()
    model_type = attr.ib()
    model_id = attr.ib()
    model_offset = attr.ib()
    address_offset = attr.ib()
    repeating_block_reference = attr.ib(default=None)
    parameter_uuid_finder = attr.ib(default=None)
    sunspec_id = attr.ib(default=None)
    is_table = attr.ib(default=False)
    fixed_block_reference = attr.ib(default=None, type=epcpm.sunspecmodel.FixedBlock)

    def gen(self):
        # TODO: CAMPid 07548795421667967542697543743987

        scale_factor_from_uuid = epcpm.pm_helper.build_uuid_scale_factor_dict(
            points=self.wrapped.children,
            parameter_uuid_finder=self.parameter_uuid_finder,
        )

        rows = []

        points = list(self.wrapped.children)

        if self.add_padding:
            point = epcpm.sunspecmodel.DataPoint(
                type_uuid=self.padding_type.uuid,
                block_offset=(
                    self.wrapped.children[-1].block_offset
                    + self.wrapped.children[-1].size
                ),
                size=self.padding_type.value,
            )
            # TODO: ack!  just to get the address offset calculated but
            #       not calling append_child() because i don't want to shove
            #       this into the model.  :[
            point.tree_parent = self.wrapped
            points.append(point)

        summed_increments = 0

        for child in points:
            builder = builders.wrap(
                wrapped=child,
                add_padding=self.add_padding,
                padding_type=self.padding_type,
                model_type=self.model_type,
                scale_factor_from_uuid=scale_factor_from_uuid,
                parameter_uuid_finder=self.parameter_uuid_finder,
                model_id=self.model_id,
                model_offset=self.model_offset,
                is_table=self.is_table,
                repeating_block_reference=self.repeating_block_reference,
                address_offset=self.address_offset + summed_increments,
                sunspec_id=self.sunspec_id,
            )
            built_rows, address_offset_increment = builder.gen()
            summed_increments += address_offset_increment
            rows.append(built_rows)

        return rows, summed_increments


@builders(epcpm.sunspecmodel.TableBlock)
@attr.s
class TableBlock:
    """Excel spreadsheet generator for the SunSpec TableBlock class."""

    wrapped = attr.ib()
    add_padding = attr.ib()
    padding_type = attr.ib()
    model_type = attr.ib()
    model_id = attr.ib()
    model_offset = attr.ib()
    address_offset = attr.ib()
    repeating_block_reference = attr.ib(default=None)
    parameter_uuid_finder = attr.ib(default=None)
    sunspec_id = attr.ib(default=None)
    is_table = attr.ib(default=False)
    fixed_block_reference = attr.ib(default=None, type=epcpm.sunspecmodel.FixedBlock)

    def gen(self) -> typing.List[typing.List[Fields], int]:
        """
        Excel spreadsheet generator for the SunSpec TableBlock class.

        Returns:
            list of Fields: int: output list of Fields, address length of block
        """
        rows = []

        groups = list(self.wrapped.children)

        summed_increments = 0

        scale_factor_from_uuid = epcpm.pm_helper.build_uuid_scale_factor_dict(
            points=self.fixed_block_reference.children,
            parameter_uuid_finder=self.parameter_uuid_finder,
        )

        for child in groups:
            builder = builders.wrap(
                wrapped=child,
                add_padding=self.add_padding,
                padding_type=self.padding_type,
                model_type=self.model_type,
                scale_factor_from_uuid=scale_factor_from_uuid,
                parameter_uuid_finder=self.parameter_uuid_finder,
                model_id=self.model_id,
                model_offset=self.model_offset,
                is_table=self.is_table,
                repeating_block_reference=self.repeating_block_reference,
                address_offset=self.address_offset + summed_increments,
                sunspec_id=self.sunspec_id,
            )
            built_rows, address_offset_increment = builder.gen()
            summed_increments += address_offset_increment
            rows.extend(built_rows)

        return rows, summed_increments


@builders(epcpm.sunspecmodel.TableGroup)
@attr.s
class TableGroup:
    """Excel spreadsheet generator for the SunSpec TableGroup class."""

    wrapped = attr.ib()
    add_padding = attr.ib()
    padding_type = attr.ib()
    model_type = attr.ib()
    scale_factor_from_uuid = attr.ib(
        type=typing.Dict[uuid.UUID, epcpm.sunspecmodel.DataPoint]
    )
    model_id = attr.ib()
    model_offset = attr.ib()
    address_offset = attr.ib()
    repeating_block_reference = attr.ib(default=None)
    parameter_uuid_finder = attr.ib(default=None)
    sunspec_id = attr.ib(default=None)
    is_table = attr.ib(default=False)

    def gen(self) -> typing.List[typing.List[Fields], int]:
        """
        Excel spreadsheet generator for the SunSpec TableGroup class.

        Returns:
            list of Fields: int: output list of Fields, address length of block
        """
        rows = []

        points = list(self.wrapped.children)

        summed_increments = 0

        for child in points:
            builder = builders.wrap(
                wrapped=child,
                add_padding=self.add_padding,
                padding_type=self.padding_type,
                model_type=self.model_type,
                scale_factor_from_uuid=self.scale_factor_from_uuid,
                parameter_uuid_finder=self.parameter_uuid_finder,
                model_id=self.model_id,
                model_offset=self.model_offset,
                is_table=self.is_table,
                repeating_block_reference=self.repeating_block_reference,
                address_offset=self.address_offset + summed_increments,
                sunspec_id=self.sunspec_id,
            )
            built_rows, address_offset_increment = builder.gen()
            summed_increments += address_offset_increment

            if isinstance(child, epcpm.sunspecmodel.TableGroup):
                rows.extend(built_rows)
            else:
                rows.append(built_rows)

        return rows, summed_increments


@builders(epcpm.sunspecmodel.DataPointBitfield)
@attr.s
class DataPointBitfield:
    wrapped = attr.ib()
    add_padding = attr.ib()
    padding_type = attr.ib()
    model_type = attr.ib()
    scale_factor_from_uuid = attr.ib()
    parameter_uuid_finder = attr.ib()
    model_id = attr.ib()
    model_offset = attr.ib()
    is_table = attr.ib()
    repeating_block_reference = attr.ib()
    address_offset = attr.ib()
    sunspec_id = attr.ib(default=None)

    def gen(self):
        row = Fields()
        row.address_offset = self.address_offset
        row.modbus_address = self.model_offset + self.address_offset
        row.field_type = self.model_type

        for name, field in attr.asdict(bitfield_fields).items():
            if field is None:
                continue

            setattr(row, name, getattr(self.wrapped, field.name))

        if row.type is not None:
            row.type = self.parameter_uuid_finder(row.type).name

        if self.wrapped.parameter_uuid is not None:
            parameter = self.parameter_uuid_finder(self.wrapped.parameter_uuid)

            row.label = parameter.name
            row.name = parameter.abbreviation
            row.notes = "" if parameter.notes is None else parameter.notes
            row.notes = f"{row.notes}  <uuid:{parameter.uuid}>".strip()

            if row.units is None:
                row.units = parameter.units
            row.description = parameter.comment
            row.read_write = "R" if parameter.read_only else "RW"

        (
            get_out,
            set_out,
        ) = epcpm.sunspectointerface.sunspec_interface_generation_for_data_point_bitfield(
            parameter, self.sunspec_id
        )
        row.get = get_out
        row.set = set_out

        return row, row.size


@enumeration_builders(epcpm.sunspecmodel.DataPointBitfield)
@attr.s
class DataPointBitfield:
    wrapped = attr.ib()
    point = attr.ib()
    parameter_uuid_finder = attr.ib(default=None)

    def gen(self):
        rows = []

        # 16 bits per register
        total_bit_count = self.wrapped.size * 16
        decimal_digits = len(str(total_bit_count - 1))

        enumerators_by_bit = {
            enumerator.bit_offset
            + i: [enumerator, i if enumerator.bit_length > 1 else None]
            for enumerator in self.wrapped.children
            for i in range(enumerator.bit_length)
        }

        member_parameter = self.parameter_uuid_finder(
            self.wrapped.parameter_uuid,
        )

        for bit in range(total_bit_count):
            enumerator, index = enumerators_by_bit.get(bit, [None, None])

            if enumerator is None:
                padded_bit_string = f"{bit:0{decimal_digits}}"
                row = Fields(
                    field_type=f"bitfield{total_bit_count}",
                    value=bit,
                    applicable_point=member_parameter.abbreviation,
                    name=f"Rsvd{padded_bit_string}",
                    label=f"Reserved - {padded_bit_string}",
                )
            else:
                builder = enumerator_builders.wrap(
                    wrapped=enumerator,
                    point=self.point,
                    parameter_uuid_finder=self.parameter_uuid_finder,
                )
                row = builder.gen()

                if index is not None:
                    padded_index_string = f"{index:0{decimal_digits}}"

                    row = attr.evolve(
                        row,
                        name=f"{row.name}{padded_index_string}",
                        label=f"{row.label} - {padded_index_string}",
                        value=row.value + index,
                    )

            rows.append(row)

        return rows


@enumerator_builders(epcpm.sunspecmodel.DataPointBitfieldMember)
@attr.s
class DataPointBitfieldMember:
    wrapped = attr.ib()
    point = attr.ib()
    parameter_uuid_finder = attr.ib(default=None)

    # TODO: CAMPid 07397546759269756456100183066795496952476951653
    def gen(self):
        field_parameter = self.parameter_uuid_finder(self.point.parameter_uuid)
        member_parameter = self.parameter_uuid_finder(self.wrapped.parameter_uuid)

        length = self.point.size * 16

        row = Fields(
            field_type=f"bitfield{length}",
            value=self.wrapped.bit_offset,
            applicable_point=field_parameter.abbreviation,
            name=member_parameter.abbreviation,
            label=member_parameter.name,
            description=member_parameter.comment,
            notes=member_parameter.notes,
        )

        return row


@enumeration_builders(epcpm.sunspecmodel.HeaderBlock)
@enumeration_builders(epcpm.sunspecmodel.FixedBlock)
@enumeration_builders(epcpm.sunspecmodel.TableRepeatingBlockReference)
@attr.s
class GenericEnumeratorBuilder:
    wrapped = attr.ib()
    parameter_uuid_finder = attr.ib(default=None)

    def gen(self):
        rows = []

        for child in self.wrapped.children:
            builder = enumeration_builders.wrap(
                wrapped=child,
                point=child,
                parameter_uuid_finder=self.parameter_uuid_finder,
            )

            new_rows = builder.gen()

            if len(new_rows) > 0:
                rows.extend(new_rows)
                rows.append(Fields())

        return rows


@enumeration_builders(epcpm.sunspecmodel.TableBlock)
@attr.s
class TableBlockEnumeratorBuilder:
    """Excel spreadsheet generator for the SunSpec TableBlock class enumerators."""

    wrapped = attr.ib()
    parameter_uuid_finder = attr.ib(default=None)

    def gen(self) -> typing.List[Fields]:
        """
        Excel spreadsheet generator for the SunSpec TableBlock class enumerators.

        Returns:
            list of Fields: output enumerators
        """
        rows = []

        for child in self.wrapped.children:
            builder = enumeration_builders.wrap(
                wrapped=child,
                parameter_uuid_finder=self.parameter_uuid_finder,
            )

            new_rows = builder.gen()

            if len(new_rows) > 0:
                rows.extend(new_rows)
                rows.append(Fields())

        return rows


@enumeration_builders(epcpm.sunspecmodel.TableGroup)
@attr.s
class TableGroupEnumeratorBuilder:
    """Excel spreadsheet generator for the SunSpec TableGroup class enumerators."""

    wrapped = attr.ib()
    parameter_uuid_finder = attr.ib(default=None)

    def gen(self) -> typing.List[Fields]:
        """
        Excel spreadsheet generator for the SunSpec TableGroup class enumerators.

        Returns:
            list of Fields: output enumerators
        """
        rows = []

        for child in self.wrapped.children:
            if isinstance(child, epcpm.sunspecmodel.TableGroup):
                builder = enumeration_builders.wrap(
                    wrapped=child,
                    parameter_uuid_finder=self.parameter_uuid_finder,
                )
            else:
                builder = enumeration_builders.wrap(
                    wrapped=child,
                    point=child,
                    parameter_uuid_finder=self.parameter_uuid_finder,
                )

            new_rows = builder.gen()

            if len(new_rows) > 0:
                rows.extend(new_rows)
                rows.append(Fields())

        return rows


@enumeration_builders(
    epcpm.sunspecmodel.TableRepeatingBlockReferenceDataPointReference,
)
@attr.s
class TableRepeatingBlockReferenceDataPointReferenceEnumerationBuilder:
    wrapped = attr.ib()
    point = attr.ib()
    parameter_uuid_finder = attr.ib(default=None)

    def gen(self):
        builder = enumeration_builders.wrap(
            wrapped=self.wrapped.original,
            point=self.point.original,
            parameter_uuid_finder=self.parameter_uuid_finder,
        )

        return builder.gen()


@builders(epcpm.sunspecmodel.TableRepeatingBlockReference)
@attr.s
class TableRepeatingBlockReference:
    wrapped = attr.ib()
    add_padding = attr.ib()
    padding_type = attr.ib()
    model_type = attr.ib()
    model_id = attr.ib()
    model_offset = attr.ib()
    address_offset = attr.ib()
    parameter_uuid_finder = attr.ib(default=None)
    sunspec_id = attr.ib(default=None)
    is_table = attr.ib(default=False)
    fixed_block_reference = attr.ib(default=None, type=epcpm.sunspecmodel.FixedBlock)

    def gen(self):
        builder = builders.wrap(
            wrapped=self.wrapped.original,
            model_type=self.model_type,
            parameter_uuid_finder=self.parameter_uuid_finder,
            add_padding=self.add_padding,
            padding_type=self.padding_type,
            model_id=self.model_id,
            model_offset=self.model_offset,
            is_table=self.is_table,
            repeating_block_reference=self.wrapped,
            address_offset=self.address_offset,
            sunspec_id=self.sunspec_id,
        )

        return builder.gen()


@builders(epcpm.sunspecmodel.DataPoint)
@attr.s
class Point:
    wrapped = attr.ib()
    add_padding = attr.ib()
    padding_type = attr.ib()
    scale_factor_from_uuid = attr.ib()
    model_type = attr.ib()
    model_id = attr.ib()
    model_offset = attr.ib()
    is_table = attr.ib()
    address_offset = attr.ib()
    repeating_block_reference = attr.ib()
    parameter_uuid_finder = attr.ib(default=None)
    sunspec_id = attr.ib(default=None)

    # TODO: CAMPid 07397546759269756456100183066795496952476951653
    def gen(self):
        row = Fields()
        row.address_offset = self.address_offset
        row.modbus_address = self.model_offset + self.address_offset

        for name, field in attr.asdict(point_fields).items():
            if field is None:
                continue

            setattr(row, name, getattr(self.wrapped, field.name))

        if row.type is not None:
            row.type = self.parameter_uuid_finder(row.type).name

        row.mandatory = "M" if self.wrapped.mandatory else "O"

        if self.wrapped.parameter_uuid is not None:
            if self.repeating_block_reference is not None:
                target = self.parameter_uuid_finder(
                    self.wrapped.parameter_uuid
                ).original
                is_array_element = isinstance(
                    target,
                    epyqlib.pm.parametermodel.ArrayParameterElement,
                )
                if is_array_element:
                    target = target.tree_parent.children[0]

                references = [
                    child
                    for child in self.repeating_block_reference.children
                    if target.uuid == child.parameter_uuid
                ]
                if len(references) > 0:
                    (reference,) = references

                    if reference.factor_uuid is not None:
                        row.scale_factor = self.parameter_uuid_finder(
                            self.parameter_uuid_finder(
                                reference.factor_uuid
                            ).parameter_uuid
                        ).abbreviation
            else:
                if row.scale_factor is not None:
                    row.scale_factor = self.parameter_uuid_finder(
                        self.scale_factor_from_uuid[row.scale_factor].parameter_uuid
                    ).abbreviation

            parameter = self.parameter_uuid_finder(self.wrapped.parameter_uuid)

            row.label = parameter.name
            row.name = parameter.abbreviation
            row.notes = "" if parameter.notes is None else parameter.notes
            to_tree = list(
                itertools.takewhile(
                    lambda node: not isinstance(node, epyqlib.pm.parametermodel.Table),
                    parameter.ancestors(),
                )
            )
            tree = to_tree[-1]
            if isinstance(tree, epyqlib.pm.parametermodel.TableGroupElement):
                # naturally sorted by traversal...  hopefully
                all_of_them = tree.nodes_by_filter(
                    filter=lambda node: node.original == parameter.original,
                    collection=[],
                )
                uuids_string = " ".join(
                    (f"<uuid:{parameter.uuid}>" for parameter in all_of_them),
                )
                row.notes = f"{row.notes}  {uuids_string}".strip()
            else:
                row.notes = f"{row.notes}  <uuid:{parameter.uuid}>".strip()

            if row.units is None:
                row.units = parameter.units
            row.description = parameter.comment
            row.read_write = "R" if parameter.read_only else "RW"

            (
                get_out,
                set_out,
                _,
            ) = epcpm.sunspectointerface.sunspec_interface_generation_for_data_point(
                parameter,
                self.sunspec_id,
                self.model_id,
                self.is_table,
                self.wrapped.not_implemented,
                row.scale_factor,
                row.type,
            )
            row.get = get_out
            row.set = set_out

        row.field_type = self.model_type

        if self.parameter_uuid_finder(self.wrapped.type_uuid).name == "pad":
            row.name = "Pad"
            row.description = "Force even alignment"
            row.read_write = "R"
            row.mandatory = "O"

        return row, row.size
