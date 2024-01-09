import collections
import pathlib
import openpyxl

import epyqlib.tests.test_attrsmodel

import epcpm.project
import epcpm.anomalymodel
import epcpm.anomaliestoxlsx
import epcpm.pm_helper

# See file COPYING in this source tree
__copyright__ = "Copyright 2024, EPC Power Corp."
__license__ = "GPLv2+"


this = pathlib.Path(__file__).resolve()
here = this.parent


def test_anomaly_model():
    """
    Tests that the project with anomaly information is loaded correctly, and that anomaly models are found in the loaded project.
    """

    # Expected number of certain objects in the project
    expected_counts = {
        epcpm.anomalymodel.AnomalyTable: 2,
        epcpm.anomalymodel.Anomaly: 3,
        epcpm.anomalymodel.AnomalySource: 1,
    }

    # Recursively traverse tree and list types of elements in sequence
    def count_types(sequence):
        def check_children(node):
            for child in node.children:
                counts[type(child)] += 1
                if child.children:
                    check_children(child)

        counts = collections.defaultdict(int)
        check_children(sequence)

        return counts

    # Load model
    project = epcpm.project.loadp(here / "project" / "project.pmp")
    anomaly_tables = project.models.anomalies.root

    # Check that counts match
    assert count_types(anomaly_tables) == expected_counts


def test_anomalies_to_spreadsheet():
    """
    Tests that spreadsheet file generated from the anomaly model is correct
    """
    filename = "test_anomalies.xlsx"

    # Load model
    project = epcpm.project.loadp(here / "project" / "project.pmp")
    anomaly_model = project.models.anomalies
    parameters_model = project.models.parameters

    # Generate workbook
    column_filter = epcpm.pm_helper.attr_fill(epcpm.anomaliestoxlsx.Fields, True)
    builder = epcpm.anomaliestoxlsx.builders.wrap(
        wrapped=anomaly_model.root,
        column_filter=column_filter,
        parameter_uuid_finder=parameters_model.node_from_uuid,
        response_levels=anomaly_model.list_selection_roots["anomaly_response_levels"],
        trigger_types=anomaly_model.list_selection_roots["anomaly_trigger_types"],
        skip_output=False,
    )
    workbook = builder.gen()

    # Check names of the sheets
    assert workbook.sheetnames == ["Anomalies", "Response levels", "Trigger types"]

    ws = workbook["Anomalies"]

    # Check dimensions of the anomalies sheet
    assert ws.min_row == 1
    assert ws.max_row == 4
    assert ws.min_column == 1
    assert ws.max_column == 9

    tables = [cell.value for cell in ws["A"]]
    names = [cell.value for cell in ws["B"]]
    codes = [cell.value for cell in ws["C"]]

    # Check that certain cells exist
    assert "Anomaly Table 1" in tables
    assert "Anomaly 3" in names
    assert all(code in codes for code in [10, 20, 30])

    # Save and test that loading works to detect illegal formatting, etc.
    workbook.save(filename)
    workbook.close()
    new_file = openpyxl.load_workbook(filename)
    new_file.close()
