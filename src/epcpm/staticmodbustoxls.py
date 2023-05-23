from __future__ import (
    annotations,
)  # See PEP 563, check to remove in future Python version higher than 3.7
import attr
import openpyxl
import pathlib
import typing

import epcpm
import epcpm.pm_helper
import epcpm.staticmodbusmodel
import epyqlib.attrsmodel
import epyqlib.pm.parametermodel
import epyqlib.utils.general


builders = epyqlib.utils.general.TypeMap()

@attr.s
class Fields(epcpm.pm_helper.FieldsInterface):
    """The fields defined for a given row in the output XLS file."""

    modbus_address = attr.ib(default=None, type=typing.Union[str, bool, int])
    name = attr.ib(default=None, type=typing.Union[str, bool])
    label = attr.ib(default=None, type=typing.Union[str, bool])
    size = attr.ib(default=None, type=typing.Union[str, bool, int])
    type = attr.ib(default=None, type=typing.Union[str, bool])
    units = attr.ib(default=None, type=typing.Union[str, bool])
    read_write = attr.ib(default=None, type=typing.Union[str, bool])
    description = attr.ib(default=None, type=typing.Union[str, bool])
    parameter_uses_interface_item = attr.ib(default=False, type=typing.Union[str, bool])
    access_level = attr.ib(default=None, type=typing.Union[str, bool, int])
    min = attr.ib(default=None, type=typing.Union[str, bool, int])
    max = attr.ib(default=None, type=typing.Union[str, bool, int])
    bit_offset = attr.ib(default=None, type=typing.Union[str, bool, int])
    scale_factor = attr.ib(default=None, type=typing.Union[str, bool])
    enumerator = attr.ib(default=None, type=typing.Union[str, bool, int])


field_names = Fields(
    modbus_address="Modbus Address",
    name="Name",
    label="Label",
    size="Size",
    type="Type",
    units="Units",
    read_write="R/W",
    description="Description",
    parameter_uses_interface_item="Implemented",
    access_level="Access Level",
    min="Min",
    max="Max",
    bit_offset="Bit Offset",
    scale_factor="SF",
    enumerator="Enumerator"
)

def build_uuid_scale_factor_dict(points, parameter_uuid_finder) -> dict:
    """
    Function that creates a dict of enumerator dicts found by using
    parameter_uuid_finder

    Args:
        points (list): a list of parameter dicts
        parameter_uuid_finder (function): function to identify a node given a UUID

    Returns:
        dict: a dict where the key is the parameter UUID and the value is the
        parameter dict associated with the UUID
    """
    # TODO: CAMPid 45002738594281495565841631423784
    scale_factor_from_uuid = {}
    for point in points:
        if point.type_uuid is None:
            continue

        type_node = parameter_uuid_finder(point.type_uuid)

        if type_node is None:
            continue

        if type_node.name == "staticmodbussf":
            scale_factor_from_uuid[point.uuid] = point

    return scale_factor_from_uuid


def export(
    path: pathlib.Path,
    staticmodbus_model: epyqlib.attrsmodel.Model,
    parameters_model: epyqlib.attrsmodel.Model,
    column_filter: epcpm.pm_helper.FieldsInterface = None,
    skip_output: bool = False,
) -> None:
    """
    Generate the static modbus model data Excel .xls file.
    Args:
        path: path and filename for .xls file
        staticmodbus_model: static modbus model
        parameters_model: parameters model
        column_filter: columns to be output to .xls file
        skip_output: skip output of the generated files
    Returns:
    """
    if skip_output:
        return

    if column_filter is None:
        column_filter = epcpm.pm_helper.attr_fill(Fields, True)

    builder = epcpm.staticmodbustoxls.builders.wrap(
        wrapped=staticmodbus_model.root,
        parameter_uuid_finder=staticmodbus_model.node_from_uuid,
        parameter_model=parameters_model,
        column_filter=column_filter,
    )

    workbook = builder.gen()

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


@builders(epcpm.staticmodbusmodel.Root)
@attr.s
class Root:
    """Excel spreadsheet generator for the static modbus Root class."""

    wrapped = attr.ib(type=epcpm.staticmodbusmodel.Root)
    column_filter = attr.ib(type=Fields)
    parameter_uuid_finder = attr.ib(default=None, type=typing.Callable)
    parameter_model = attr.ib(default=None, type=epyqlib.attrsmodel.Model)

    def gen(self) -> openpyxl.workbook.workbook.Workbook:
        """
        Excel spreadsheet generator for the static modbus root class.
        Returns:
            workbook: generated Excel workbook
        """
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
        modbus_worksheet = workbook.create_sheet("Static Modbus Data")
        enumeration_worksheet = workbook.create_sheet("Enumerations")
        modbus_worksheet.append(field_names.as_filtered_tuple(self.column_filter))
        enumeration_worksheet.append(["Enumerator", "Name", "Value"])

        scale_factor_from_uuid = build_uuid_scale_factor_dict(
            points=self.wrapped.children,
            parameter_uuid_finder=self.parameter_uuid_finder
        )
        enumerations = self.collect_enumerations()
        enumerators_from_uuid = {}
        for enumeration in enumerations:
            enumerators_from_uuid[enumeration.uuid] = enumeration

        # Collects the cell coordinates of each new enumerator as they're appended
        enumeration_cell_coordinates = {}
        ENUMERATION_ENUMERATOR_COLUMN = "#Enumerations!A"
        for enumeration in enumerations:
            for child in enumeration.children:
                enumeration_worksheet.append([enumeration.name, child.name, child.value])
                if enumeration.name not in enumeration_cell_coordinates.keys():
                    enumeration_cell_coordinates[enumeration.name] = (
                        ENUMERATION_ENUMERATOR_COLUMN + f"{enumeration_worksheet.max_row}"
                    )

        for member in self.wrapped.children:
            builder = builders.wrap(
                wrapped=member,
                parameter_uuid_finder=self.parameter_uuid_finder,
                scale_factor_from_uuid=scale_factor_from_uuid,
                enumerators_from_uuid=enumerators_from_uuid
            )
            rows = builder.gen()

            # Appends the rows of static modbus data
            MODBUS_ENUMERATOR_COLUMN = "O"
            for row in rows:
                modbus_worksheet.append(row.as_filtered_tuple(self.column_filter))
                # Checks if there's an enumerator and hyperlinks it to the right row in
                # the Enumerations worksheet
                if row.enumerator is not None:
                    max_row = modbus_worksheet.max_row
                    if row.enumerator in enumeration_cell_coordinates.keys():
                        modbus_worksheet[MODBUS_ENUMERATOR_COLUMN +f"{max_row}"].hyperlink = (
                            enumeration_cell_coordinates[row.enumerator]
                        )

        return workbook

    def collect_enumerations(self):
        collected = []

        if self.parameter_model is None:
            return collected

        def collect(node, collected):
            is_enumeration = isinstance(
                node,
                (
                    epyqlib.pm.parametermodel.Enumeration,
                    epyqlib.pm.parametermodel.AccessLevels,
                ),
            )
            if is_enumeration:
                collected.append(node)

        self.parameter_model.root.traverse(
            call_this=collect,
            payload=collected,
            internal_nodes=True,
        )

        return collected

@builders(epcpm.staticmodbusmodel.FunctionData)
@attr.s
class FunctionData:
    """Excel spreadsheet generator for the static modbus FunctionData class."""

    wrapped = attr.ib(type=epcpm.staticmodbusmodel.FunctionData)
    parameter_uuid_finder = attr.ib(type=typing.Callable)
    scale_factor_from_uuid = attr.ib(type=typing.Callable)
    enumerators_from_uuid = attr.ib(type=typing.Callable)

    def gen(self) -> typing.List[Fields]:
        """
        Excel spreadsheet generator for the static modbus FunctionData class.
        Returns:
            list: single Fields row of FunctionData
        """
        type_node = self.parameter_uuid_finder(self.wrapped.type_uuid)
        row = Fields()
        row.modbus_address = self.wrapped.address
        row.type = type_node.name
        row.size = self.wrapped.size
        row.units = self.wrapped.units

        if self.wrapped.parameter_uuid is not None:
            parameter = self.parameter_uuid_finder(self.wrapped.parameter_uuid)
            row.label = parameter.name
            row.name = parameter.abbreviation
            row.description = parameter.comment
            row.read_write = "R" if parameter.read_only else "RW"
            row.min = parameter.minimum
            row.max = parameter.maximum
            uses_interface_item = (
                isinstance(parameter, epyqlib.pm.parametermodel.Parameter)
                and parameter.uses_interface_item()
            )
            row.parameter_uses_interface_item = uses_interface_item
            if parameter.access_level_uuid is not None:
                access_level = self.parameter_uuid_finder(parameter.access_level_uuid)
                row.access_level = access_level.value

            if self.wrapped.factor_uuid is not None:
                row.scale_factor = self.parameter_uuid_finder(
                    self.scale_factor_from_uuid[self.wrapped.factor_uuid].parameter_uuid
                ).abbreviation

            if self.wrapped.enumeration_uuid is not None:
                row.enumerator = self.parameter_uuid_finder(
                    self.enumerators_from_uuid[self.wrapped.enumeration_uuid].uuid
                ).name

        if type_node.name == "pad":
            row.name = "Pad"
            row.description = "Force even alignment"
            row.read_write = "R"

        return [row]


@builders(epcpm.staticmodbusmodel.FunctionDataBitfield)
@attr.s
class FunctionDataBitfield:
    """Excel spreadsheet generator for the static modbus FunctionDataBitfield class."""

    wrapped = attr.ib(type=epcpm.staticmodbusmodel.FunctionDataBitfield)
    parameter_uuid_finder = attr.ib(type=typing.Callable)
    scale_factor_from_uuid = attr.ib(type=typing.Callable)
    enumerators_from_uuid = attr.ib(type=typing.Callable)

    def gen(self) -> typing.List[Fields]:
        """
        Excel spreadsheet generator for the static modbus FunctionDataBitfield class.
        Call to generate the FunctionDataBitfieldMember children.
        Returns:
            list: list of Fields rows for FunctionDataBitfield and FunctionDataBitfieldMember
        """
        rows = []
        row = Fields()
        parameter = self.parameter_uuid_finder(self.wrapped.parameter_uuid)
        row.modbus_address = self.wrapped.address
        row.size = self.wrapped.size
        row.label = parameter.name
        row.name = parameter.abbreviation
        row.description = parameter.comment
        row.read_write = "R" if parameter.read_only else "RW"
        rows.append(row)

        for member in self.wrapped.children:
            builder = builders.wrap(
                wrapped=member,
                parameter_uuid_finder=self.parameter_uuid_finder,
            )
            member_row = builder.gen()
            member_row.modbus_address = (
                self.wrapped.address
            )
            rows.append(member_row)

        return rows


@builders(epcpm.staticmodbusmodel.FunctionDataBitfieldMember)
@attr.s
class FunctionDataBitfieldMember:
    """Excel spreadsheet generator for the static modbus FunctionDataBitfieldMember class."""

    wrapped = attr.ib(type=epcpm.staticmodbusmodel.FunctionDataBitfield)
    parameter_uuid_finder = attr.ib(type=typing.Callable)

    def gen(self) -> Fields:
        """
        Excel spreadsheet generator for the static modbus FunctionDataBitfieldMember class.
        Returns:
            row: Fields row for a FunctionDataBitfieldMember
        """
        row = Fields()
        type_node = self.parameter_uuid_finder(self.wrapped.type_uuid)
        row.type = type_node.name
        row.size = self.wrapped.bit_length
        row.bit_offset = self.wrapped.bit_offset
        if self.wrapped.parameter_uuid is not None:
            parameter = self.parameter_uuid_finder(self.wrapped.parameter_uuid)
            row.label = parameter.name
            row.name = parameter.abbreviation
            row.description = parameter.comment
            row.read_write = "R" if parameter.read_only else "RW"
            uses_interface_item = (
                isinstance(parameter, epyqlib.pm.parametermodel.Parameter)
                and parameter.uses_interface_item()
            )
            row.parameter_uses_interface_item = uses_interface_item
            if parameter.access_level_uuid is not None:
                access_level = self.parameter_uuid_finder(parameter.access_level_uuid)
                row.access_level = access_level.value

        return row