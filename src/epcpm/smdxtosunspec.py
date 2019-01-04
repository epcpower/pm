import collections
import contextlib
import os

import attr
import epyqlib.pm.parametermodel
import openpyxl
import sunspec.core.device

import epcpm.sunspecmodel


def epc_point_from_pysunspec_point(point, parameter_model, parameter_uuid, scale_factors=None):
    if scale_factors is not None and point.point_type.sf is not None:
        scale_factor_uuid = scale_factors[point.point_type.sf].uuid
    else:
        scale_factor_uuid = None

    sunspec_type_uuid = point.point_type.type
    if sunspec_type_uuid is not None:
        root = parameter_model.list_selection_roots['sunspec types']
        sunspec_type_uuid = root.child_by_name(sunspec_type_uuid).uuid

    return epcpm.sunspecmodel.DataPoint(
        factor_uuid=scale_factor_uuid,
        parameter_uuid=parameter_uuid,
        type_uuid=sunspec_type_uuid,
        # enumeration_uuid=,
        block_offset=point.point_type.offset,
        size=point.point_type.len,
        mandatory=point.point_type.mandatory == 'true',
        # uuid=,
    )


def epc_parameter_from_pysunspec_point(point):
    parameter = epyqlib.pm.parametermodel.Parameter(
        name=point.point_type.label,
        abbreviation=point.point_type.id,
        notes=point.point_type.notes,
        units=point.point_type.units,
        comment=point.point_type.description,
        read_only='w' not in point.point_type.access,
    )

    return parameter


def import_models(*model_ids, parameter_model, paths):
    return [
        import_model(model_id=id, parameter_model=parameter_model, paths=paths)
        for id in model_ids
    ]


@contextlib.contextmanager
def fresh_smdx_path(*paths):
    original_pathlist = sunspec.core.device.file_pathlist
    sunspec.core.device.file_pathlist = sunspec.core.util.PathList()

    for path in paths:
        sunspec.core.device.file_pathlist.add(os.fspath(path))

    try:
        yield sunspec.core.device.file_pathlist
    finally:
        sunspec.core.device.file_pathlist = original_pathlist


def none_to_empty_string(value):
    if value is None:
        return ''

    return value


@attr.s(frozen=True)
class Symbol:
    value = attr.ib(converter=int)
    id = attr.ib()
    label = attr.ib()
    type = attr.ib()
    description = attr.ib(converter=none_to_empty_string)
    notes = attr.ib(converter=none_to_empty_string)

    @classmethod
    def from_sunspec(cls, symbol, type_):
        return cls(
            description=symbol.description,
            id=symbol.id,
            label=symbol.label,
            notes=symbol.notes,
            value=symbol.value,
            type=type_,
        )


def import_model(model_id, parameter_model, paths=()):
    model = sunspec.core.device.Model(mid=model_id)
    if len(paths) == 0:
        model.load()
    else:
        with fresh_smdx_path(*paths):
            model.load()

    imported_points = []
    scale_factors = {}

    group = epyqlib.pm.parametermodel.Group(
        name='SunSpec Model {}'.format(model.id),
    )
    parameter_model.root.append_child(group)

    our_model = epcpm.sunspecmodel.Model(
        id=model.id,
        length=model.len,
    )

    types = parameter_model.list_selection_roots['sunspec types']
    parameters = our_model.children[0].add_data_points(
        model_id=model.model_type.label,
        uint16_uuid=types.child_by_name('uint16').uuid,
    )

    for parameter in parameters:
        group.append_child(parameter)

    for name, point in model.points_sf.items():
        parameter = epc_parameter_from_pysunspec_point(point=point)
        group.append_child(parameter)
        epc_point = epc_point_from_pysunspec_point(
            point=point,
            parameter_uuid=parameter.uuid,
            parameter_model=parameter_model,
        )
        scale_factors[name] = epc_point

    enumerations = collections.defaultdict(list)

    for point in model.points_list:
        parameter = epc_parameter_from_pysunspec_point(point=point)
        group.append_child(parameter)
        epc_point = epc_point_from_pysunspec_point(
            point=point,
            parameter_model=parameter_model,
            parameter_uuid=parameter.uuid,
            scale_factors=scale_factors,
        )

        imported_points.append(epc_point)

        if point.point_type.type.startswith(('enum', 'bitfield')):
            enumeration = tuple(sorted(
                Symbol.from_sunspec(symbol=symbol, type_=point.point_type.type)
                for symbol in point.point_type.symbols
            ))
            enumerations[enumeration].append(epc_point)

    enumerations_root = parameter_model.list_selection_roots['enumerations']
    for enumeration, points in enumerations.items():
        parameter = parameter_model.node_from_uuid(points[0].parameter_uuid)
        # TODO: just using the first point?  hmm
        epc_enumeration = epyqlib.pm.parametermodel.Enumeration(
            name='SunSpec{}'.format(parameter.name),
        )

        points[0].enumeration_uuid = epc_enumeration.uuid

        for symbol in enumeration:
            enumerator = epyqlib.pm.parametermodel.SunSpecEnumerator(
                name=symbol.id,
                label=symbol.label,
                description=symbol.description,
                notes=symbol.notes,
                value=symbol.value,
                type=symbol.type,
            )
            epc_enumeration.append_child(enumerator)

        enumerations_root.append_child(epc_enumeration)

    id_point = our_model.children[0].children[0]
    id_point.id = model.model_type.id
    id_point.notes = model.model_type.notes

    parameters[0].comment = model.model_type.description

    imported_points = sorted(
        imported_points + list(scale_factors.values()),
        key=lambda point: point.block_offset,
    )

    for point in imported_points:
        our_model.children[1].append_child(point)

    return our_model


@attr.s(frozen=True)
class GetSetKey:
    model = attr.ib()
    name = attr.ib()
    get_set = attr.ib()


def import_get_set(path):
    workbook = openpyxl.load_workbook(path)

    collected = {}

    for sheet in workbook.worksheets:
        try:
            model = int(sheet.title)
        except ValueError:
            continue

        iter_rows = iter(sheet.rows)

        column_indexes = {
            cell.value: i
            for i, cell in enumerate(next(iter_rows))
        }

        for row in iter_rows:
            for get_set in ('get', 'set'):
                value = row[column_indexes[get_set]].value

                if value in (None, ''):
                    continue

                name = row[column_indexes['Name']].value
                key = GetSetKey(
                    model=model,
                    name=name,
                    get_set=get_set,
                )
                collected[key] = value

    return collected
